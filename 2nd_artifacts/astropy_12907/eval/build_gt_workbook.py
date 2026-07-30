"""Build ground_truth_astropy12907_DRAFT.xlsx: answer key + GT tree +
adjudication list, every judgment row marked PROVISIONAL for author.
Template style per project convention: Calibri; bold-13 merged title;
10pt merged description; bold-11 headers; 11pt body; formulas for counts.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

TITLE = Font(name="Calibri", size=13, bold=True)
DESC = Font(name="Calibri", size=10)
HEAD = Font(name="Calibri", size=11, bold=True)
BODY = Font(name="Calibri", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
FILL = PatternFill("solid", fgColor="FFF2CC")  # provisional highlight

wb = Workbook()


def sheet_head(ws, ncols, title, desc):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(1, 1, title).font = TITLE
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(2, 2 - 1, desc)
    c.font = DESC
    c.alignment = WRAP
    ws.row_dimensions[2].height = 40


def put_row(ws, r, values, fill=False):
    for i, v in enumerate(values, 1):
        c = ws.cell(r, i, v)
        c.font = BODY
        c.alignment = WRAP
        if fill:
            c.fill = FILL


# ------------------------- Sheet 1: Answer key -------------------------
ws = wb.active
ws.title = "Answer_key"
cols = ["GT id", "Required step", "Grounded in (Ctrl-F anchor)", "Status"]
sheet_head(ws, len(cols),
           "Ground truth answer key: astropy__astropy-12907 (DRAFT)",
           "Required steps implied by the intent, each grounded in the issue "
           "text or benchmark artifacts. EVERY ROW IS PROVISIONAL: author "
           "adjudicates (accept / edit / strike) and sets Status. Legend: "
           "edit only the Status and text cells of highlighted rows; "
           "example of an accepted row: Status = ACCEPTED.")
for i, h in enumerate(cols, 1):
    ws.cell(4, i, h).font = HEAD

key = [
    ("GT-01", "Reproduce the reported incorrect behavior using the "
              "reproduction the intent supplies",
     "intent: 'Suddenly the inputs and outputs are no longer separable?' "
     "plus the three ```python blocks"),
    ("GT-02", "Localize the defect within the separability computation",
     "intent: 'does not compute separability correctly' names the failing "
     "capability, not the location; locating is implied work"),
    ("GT-03", "Characterize why nesting specifically breaks the computation",
     "intent contrasts flat vs nested outcomes across its three examples"),
    ("GT-04", "Correct the computation so nested CompoundModels yield the "
              "expected matrix",
     "reference patch: one-line _cstack fix (cright[...] = right); the "
     "intent's core obligation"),
    ("GT-05", "Verify the reported case now yields the expected output",
     "intent supplies the expected diagonal/block matrices as the oracle"),
    ("GT-06", "Preserve existing separability behavior (no regressions)",
     "benchmark PASS_TO_PASS: 13 existing tests must keep passing; implied "
     "by 'fix correctly'"),
]
r = 5
for gid, step, ground in key:
    put_row(ws, r, [gid, step, ground, "PROVISIONAL"], fill=True)
    r += 1
put_row(ws, r + 1, ["", "Answer-key steps counted:",
                    "=COUNTA(A5:A%d)" % (r - 1), ""])
for i, w in enumerate([10, 52, 52, 14], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ------------------------- Sheet 2: GT tree -------------------------
ws = wb.create_sheet("GT_tree")
cols = ["id", "parent_id", "level", "definer", "enforcer", "action",
        "resource", "spatial", "temporal", "context", "rationale",
        "Realizes", "Status"]
sheet_head(ws, len(cols),
           "Ground truth policy tree: astropy__astropy-12907 (DRAFT)",
           "Draft reference tree the structured/unstructured are scored against. Working nodes "
           "map 1:1 to answer-key steps via Realizes; structural nodes are "
           "labeled scaffolding. The arbitrate row extends the project's "
           "safety-first invariant convention to a repair intent and NEEDS "
           "ADJUDICATION. Edit highlighted cells; set Status per row.")
for i, h in enumerate(cols, 1):
    ws.cell(4, i, h).font = HEAD

AE = "assurance engine"
tree = [
    ("P0", None, "declarative", "human operator", "Orchestrator", "fix",
     "separability computation for nested CompoundModels",
     "astropy repository", "always", "production",
     "Restates the raw intent as the root declarative policy.",
     "scaffolding (root)"),
    ("P1", "P0", "definitive", AE, "Analyze", "locate",
     "code path computing separability for CompoundModels",
     "astropy/modeling/separable.py", "always", "production",
     "The intent names a failing capability, not a location; the defect "
     "must be located before it can be characterized or fixed.",
     "GT-02"),
    ("P2", "P1", "definitive", AE, "Analyze", "characterize",
     "difference in matrix assembly between flat and nested operands",
     "astropy/modeling/separable.py", "always", "production",
     "Locating the path is not enough; why nesting diverges from the flat "
     "case must be understood to plan a correct fix.",
     "GT-03"),
    ("P3", "P2", "definitive", AE, "Plan", "decide",
     "repair strategy for the characterized defect",
     "astropy/modeling/separable.py", "always", "production",
     "A fix, its verification, and regression protection are planned from "
     "the characterization before anything executes.",
     "scaffolding (plan hub)"),
    ("P4", "P0", "imperative", AE, "Execute", "run",
     "reproduction script from the intent's code blocks",
     "local repository workspace", "always", "production",
     "Running the supplied reproduction confirms the defect exists in the "
     "current code and fixes the baseline for later verification.",
     "GT-01"),
    ("P5", "P3", "imperative", AE, "Execute", "edit",
     "the defective assembly step in separable.py",
     "astropy/modeling/separable.py", "always", "production",
     "Applying the planned correction to the located step is the concrete "
     "realization of the intent's core obligation.",
     "GT-04"),
    ("P6", "P3", "imperative", AE, "Execute", "validate",
     "reported nested case against the intent's expected matrices",
     "local repository workspace", "always", "production",
     "The intent supplies the expected output; the patched code must "
     "reproduce it for the reported case.",
     "GT-05"),
    ("P7", "P3", "imperative", AE, "Execute", "run",
     "existing separability test suite",
     "astropy/modeling/tests/test_separable.py", "always", "production",
     "A correct fix must not alter previously correct behavior; the "
     "existing suite is the guard.",
     "GT-06"),
    ("P8", "P3", "definitive", AE, "Plan", "arbitrate",
     "scope of permitted changes (minimal, defect-local edits only)",
     "astropy repository", "always", "production",
     "Safety-first invariant: enforcement/repair actions must stay within "
     "the minimal scope the intent justifies. DERIVED, not traceable to "
     "intent text.",
     "safety-first invariant (derived; excluded from recall)"),
]
r = 5
for row in tree:
    put_row(ws, r, list(row) + ["PROVISIONAL"], fill=True)
    r += 1
s = r + 1
put_row(ws, s, ["", "", "", "", "", "", "", "", "", "", "",
                "total nodes:", "=COUNTA(A5:A%d)" % (r - 1)])
put_row(ws, s + 1, ["", "", "", "", "", "", "", "", "", "", "",
                    "working nodes (Realizes GT-*):",
                    '=COUNTIF(L5:L%d,"GT-*")' % (r - 1)])
widths = [6, 9, 11, 15, 11, 12, 40, 30, 10, 11, 55, 30, 13]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ------------------------- Sheet 3: Adjudication -------------------------
ws = wb.create_sheet("Adjudication")
cols = ["#", "Decision needed", "Draft recommendation", "Author's ruling"]
sheet_head(ws, len(cols),
           "Adjudication list: decisions of authors",
           "Each row is a judgment call the draft could not make. Record the "
           "ruling in the last column; rulings override every draft row in "
           "the other sheets. Example ruling format: 'ACCEPTED as drafted' "
           "or 'CHANGED: ...'.")
for i, h in enumerate(cols, 1):
    ws.cell(4, i, h).font = HEAD

adj = [
    (1, "GT-06 scope: is 'preserve existing behavior' an implied step of "
        "'fix correctly'?",
     "Yes; grounded in PASS_TO_PASS being part of resolved."),
    (2, "Are NEW regression tests (model's P4.2/P6/P6.1) inside the intent "
        "or beyond it?",
     "Beyond-intent: the user asked for a fix, not tests; the benchmark's "
     "test_patch is the oracle, not a user request. Score those nodes in "
     "precision like the tag/alert precedent."),
    (3, "Model's P7 delivery branch (commit, submit PR): relevant for "
        "precision?",
     "Beyond-intent; same precedent as item 2."),
    (4, "Model tree has imperative-under-imperative chains (P5.1 under P5; "
        "P6.1 under P6), a shape prior trees never had.",
     "Surface to co-author; request does not forbid it; no GT change."),
    (5, "Does the safety-first invariant (arbitrate) convention extend to "
        "repair-style intents? (GT row P8)",
     "Keep for cross-GT uniformity, excluded from recall as always; strike "
     "if co-author rules it enforcement-only."),
    (6, "Model's P1 Monitor node: does a one-shot repair intent warrant a "
        "continuous Monitor policy?",
     "Contestable; lean relevant-but-marginal for precision; author rules."),
    (7, "Disclosure: model tree names _cstack without repo access "
        "(training contamination from a famous bug).",
     "Disclose wherever the tree is presented; specificity and coverage "
     "may be inflated; TC/AC/EF unaffected (structural)."),
    (8, "Model P4 rationale contains a Unicode em dash (U+2014).",
     "Data stays verbatim in the JSON; escape or rewrite only at quoting "
     "time (LaTeX/notes)."),
]
r = 5
for n, q, rec in adj:
    put_row(ws, r, [n, q, rec, ""], fill=True)
    r += 1
for i, w in enumerate([4, 55, 55, 30], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save("/home/claude/ground_truth_astropy12907_DRAFT.xlsx")
print("saved")