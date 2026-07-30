"""
generate_evaluation.py

Produces the cross-domain evaluation workbook for PROVENANCE: the unaided-refiner
probe (recall/precision per model per domain) and the by-construction recoverability
(PROVENANCE vs baseline), with per-step scoring ledgers and native Excel clustered-bar
charts.

Manual scoring is authoritative. This script does not judge LLM output; it tabulates
the per-step judgments entered below and reproduces the numbers and charts from them.

Fill the AV block after running the AV probe (the code block is already filled from the
paper). Then:  pip install openpyxl  ;  python generate_evaluation.py
Output: evaluation_summary.xlsx
"""

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

# ===========================================================================
# DATA  --  per-step recovery judgments (manual, authoritative).
# A step id maps to True if the model's decomposition recovered it, else False.
# policies_emitted = number of non-root policies the model produced.
# precision = fraction of those that realize a key step (entered as scored).
# ===========================================================================

CODE = {
    "answer_key": ["GT-01", "GT-02", "GT-03", "GT-04", "GT-05",
                   "GT-06", "GT-07", "GT-08", "GT-09", "GT-10"],
    "models": {
        # Gemini recovered 5/10; missed classification, the two egress steps,
        # the scope-escape, and (collapsing access control) the permission step.
        "Gemini 2.5 Pro": {
            "recovered": {"GT-01": False, "GT-02": True,  "GT-03": True,
                          "GT-04": True,  "GT-05": False, "GT-06": False,
                          "GT-07": True,  "GT-08": False, "GT-09": True,
                          "GT-10": False},
            "policies_emitted": 7,
            "precision": 0.50,
        },
        # Copilot recovered 6/10; missed classification, the two egress steps,
        # and the scope-escape; over-expanded elsewhere (low precision).
        "Copilot (Think Deeper)": {
            "recovered": {"GT-01": False, "GT-02": True,  "GT-03": True,
                          "GT-04": True,  "GT-05": False, "GT-06": False,
                          "GT-07": True,  "GT-08": True,  "GT-09": True,
                          "GT-10": False},
            "policies_emitted": 17,
            "precision": 0.38,
        },
    },
    # Recoverability is by construction: PROVENANCE recovers all leaves, baseline none.
    "leaves": 8,
}

# AV probe scored 2026-06-21. Both models missed only GT-04 (all-lane scope), on a
# strict reading -- neither committed the cited fast-lane error, but neither
# affirmatively scoped the cap to all lanes. GT-06 (smooth) was recovered by both,
# so the AgentSpec blind spots did NOT reproduce. The cross-domain signal is
# precision: Copilot over-expands (7 of 13 policies off-scope for this intent).
# To score GT-04 leniently (lane-agnostic = all lanes), set GT-04 to True -> 6/6.
AV = {
    "answer_key": ["GT-01", "GT-02", "GT-03", "GT-04", "GT-05", "GT-06"],
    "models": {
        "Gemini 2.5 Pro": {
            "recovered": {"GT-01": True,  "GT-02": True,  "GT-03": True,
                          "GT-04": False, "GT-05": True,  "GT-06": True},
            "policies_emitted": 5,
            "precision": 1.00,
        },
        "Copilot (Think Deeper)": {
            "recovered": {"GT-01": True,  "GT-02": True,  "GT-03": True,
                          "GT-04": False, "GT-05": True,  "GT-06": True},
            "policies_emitted": 13,
            "precision": 0.46,
        },
    },
    "leaves": 5,
}

DOMAINS = {"Code execution": CODE, "Autonomous driving": AV}
OUTFILE = "evaluation_summary.xlsx"
# ===========================================================================


def recall(model):
    vals = [v for v in model["recovered"].values() if v is not None]
    if not vals:
        return None
    return sum(1 for v in vals if v) / len(model["recovered"])


def is_filled(model):
    return (model["precision"] is not None
            and all(v is not None for v in model["recovered"].values()))


wb = Workbook()

# ---- Probe sheet ---------------------------------------------------------
ws = wb.active
ws.title = "Probe"
ws.append(["Domain", "Model", "Recall", "Precision", "Steps recovered", "Policies"])
probe_rows = []
for dname, d in DOMAINS.items():
    for mname, m in d["models"].items():
        if is_filled(m):
            r = recall(m)
            recovered_n = sum(1 for v in m["recovered"].values() if v)
            ws.append([f"{dname} / {mname}", mname, round(r, 2), m["precision"],
                       f"{recovered_n}/{len(m['recovered'])}", m["policies_emitted"]])
            probe_rows.append(True)
        else:
            ws.append([f"{dname} / {mname}", mname, "TODO", "TODO", "TODO", "TODO"])
            probe_rows.append(False)

if all(probe_rows):
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Unaided refinement: recall and precision"
    chart.y_axis.title = "Score"
    chart.x_axis.title = "Domain / model"
    data = Reference(ws, min_col=3, max_col=4, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    ws.add_chart(chart, "H2")
else:
    ws["H2"] = "Chart appears once the AV probe rows are filled."

# ---- Recoverability sheet (by construction) ------------------------------
wr = wb.create_sheet("Recoverability")
wr.append(["Domain", "PROVENANCE", "Baseline", "Leaves"])
for dname, d in DOMAINS.items():
    wr.append([dname, 1.00, 0.00, d["leaves"]])
chart2 = BarChart()
chart2.type = "col"
chart2.grouping = "clustered"
chart2.title = "Recoverability by arm (TC = AC = EF)"
chart2.y_axis.title = "Score"
chart2.x_axis.title = "Domain"
data2 = Reference(wr, min_col=2, max_col=3, min_row=1, max_row=wr.max_row)
cats2 = Reference(wr, min_col=1, min_row=2, max_row=wr.max_row)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.y_axis.scaling.min = 0
chart2.y_axis.scaling.max = 1
wr.add_chart(chart2, "G2")

# ---- Per-step ledger sheets ----------------------------------------------
for dname, d in DOMAINS.items():
    wl = wb.create_sheet(f"Ledger_{dname.split()[0].lower()}")
    header = ["Step"] + list(d["models"].keys())
    wl.append(header)
    for step in d["answer_key"]:
        row = [step]
        for m in d["models"].values():
            v = m["recovered"][step]
            row.append("recovered" if v is True else "missed" if v is False else "TODO")
        wl.append(row)

wb.save(OUTFILE)
print(f"wrote {OUTFILE}")
print("Sheets: Probe (+chart), Recoverability (+chart), Ledger_code, Ledger_autonomous.")
if not all(is_filled(m) for m in AV["models"].values()):
    print("AV rows are placeholders -- fill the AV block after the probe, then rerun.")
